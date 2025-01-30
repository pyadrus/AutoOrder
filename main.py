import sqlite3

from openpyxl import load_workbook
from loguru import logger


def opening_the_database():
    """Открытие базы данных"""
    conn = sqlite3.connect('data/data.db')  # Создаем соединение с базой данных
    cursor = conn.cursor()
    return conn, cursor


def property_parsing():
    """Парсинг имущества"""
    try:
        conn, cursor = opening_the_database()
        # Открываем выбор файла Excel для чтения данных
        workbook = load_workbook(filename='data/124.xlsx')  # Загружаем выбранный файл Excel
        sheet = workbook.active

        # Создаем таблицу в базе данных, если она еще не существует, где:
        # table_number - табельный номер,
        # surname_name_patronymic - фамилия,
        # profession - профессия,
        # percent - процент
        cursor.execute('CREATE TABLE IF NOT EXISTS data (table_number, surname_name_patronymic, profession, percent)')

        cursor.execute('DELETE FROM data')
        conn.commit()  # сохранить изменения

        # Считываем данные из колонок A и H и вставляем их в базу данных
        for row in sheet.iter_rows(min_row=3, max_row=23, values_only=True):
            table_number = row[1]  # table_number - табельный номер,
            surname_name_patronymic = row[2]  # surname_name_patronymic - фамилия,
            profession = row[5]  # profession - профессия,
            percent = row[11]  # percent - процент
            logger.info(f'{table_number}, {surname_name_patronymic}, {profession}, {percent}')
            cursor.execute('INSERT INTO data VALUES (?, ?, ?, ?)',
                           (table_number, surname_name_patronymic, profession, percent,
                            ))
            # Сохраняем изменения в базе данных и закрываем соединение
        conn.commit()
        conn.close()
    except Exception as e:
        logger.exception(e)


property_parsing()
