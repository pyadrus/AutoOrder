from docxtpl import DocxTemplate
from loguru import logger
from openpyxl import load_workbook

from database import opening_the_database, get_data_from_db
from full_name_of_professions import full_name_of_professions


def supplement_for_work_wartime(data_mounts, file_dog, number_month):
    """
    Доплата за работу в военное время
    :param data_mounts: месяц, например январь или февраль
    :param file_dog: название файла шаблона, например Доплата_за_высокие_достижения_в_труде.docx
    :param number_month: номер месяца для исходных данных, например январь 01_25 или февраль 02_25
    """

    def record_data_salary_downtime_week():
        """Заполнение приказа"""
        doc = DocxTemplate(f"data/sample/{file_dog}")

        # Получаем данные из базы данных
        rows = get_data_from_db()
        table_data = prepare_table_data(rows)

        context = {
            "data_mounts": f" {data_mounts} ",
            "table_data": table_data  # Передаем данные для таблицы
        }

        doc.render(context)
        doc.save(f"data/{data_mounts}/{file_dog}")

    def prepare_table_data(rows):
        """Подготовка данных для таблицы"""
        table_data = []
        for row in rows:
            table_data.append({
                "table_number": row[0],  # Табельный номер
                "surname_name_patronymic": row[1],  # ФИО
                "profession": row[2],  # Профессия
                "percent": row[3]  # Процент
            })
        return table_data

    def property_parsing():
        """Парсинг данных"""
        try:
            conn, cursor = opening_the_database()
            # Открываем выбор файла Excel для чтения данных
            workbook = load_workbook(filename=f'data/initial_data/{number_month}/164.xlsx')  # Загружаем выбранный файл Excel
            sheet = workbook.active
            # Создаем таблицу в базе данных, если она еще не существует
            cursor.execute('CREATE TABLE IF NOT EXISTS data (table_number, surname_name_patronymic, profession, percent)')
            cursor.execute('DELETE FROM data')
            conn.commit()  # сохранить изменения

            # Считываем данные из колонок и вставляем их в базу данных
            for row in sheet.iter_rows(min_row=3, max_row=102, values_only=True):
                try:
                    # Проверяем, что строка содержит достаточно данных
                    if len(row) > 11:  # Убедимся, что в строке есть хотя бы 12 столбцов
                        table_number = row[1]  # table_number - табельный номер,
                        surname_name_patronymic = row[2]  # surname_name_patronymic - фамилия,
                        profession = row[5]  # profession - профессия,
                        percent = row[11]  # percent - процент

                        profession = full_name_of_professions.get(profession, profession)  # Получаем полное название профессии
                        print(profession)
                        # Логируем данные для отладки
                        logger.info(f'Данные для вставки: {table_number}, {surname_name_patronymic}, {profession}, {percent}')

                        # Вставляем данные в таблицу
                        cursor.execute('INSERT INTO data VALUES (?, ?, ?, ?)', (table_number, surname_name_patronymic, profession, percent))
                    else:
                        logger.warning(f"Строка содержит недостаточно данных: {row}")
                except Exception as e:
                    logger.error(f"Ошибка при обработке строки {row}: {e}")

            # Сохраняем изменения в базе данных и закрываем соединение
            conn.commit()
            conn.close()

        except FileNotFoundError:
            logger.error(f"Файл не найден: {number_month}")
            return

        except Exception as e:
            logger.exception(e)

    property_parsing()
    record_data_salary_downtime_week()
