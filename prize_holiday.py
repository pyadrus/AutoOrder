from docxtpl import DocxTemplate
from loguru import logger
from openpyxl import load_workbook
import os
from database import opening_the_database, get_data_from_db
from full_name_of_professions import full_name_of_professions


def prize_holiday(name_month, data_mounts, file_dog, year="2025"):
    """
    Премия к празднику
    :param name_month: название месяца, например "Январь"
    :param data_mounts: месяц, например "01" или "08"
    :param file_dog: название файла шаблона, например "Доплата_до_МРОТ.docx"
    :param year: год (по умолчанию 2025)
    """

    def record_data_salary_downtime_week():
        """Заполнение приказа"""
        doc = DocxTemplate(f"data/sample/{file_dog}")

        # Получаем данные из базы данных
        rows = get_data_from_db()
        table_data = prepare_table_data(rows)

        context = {
            "data_mounts": f" {name_month} ",
            "table_data": table_data  # Передаем данные для таблицы
        }

        output_dir = f"data/{year}/output/{data_mounts}"
        os.makedirs(output_dir, exist_ok=True)  # создаем папку, если её нет

        output_path = os.path.join(output_dir, file_dog)
        doc.render(context)
        doc.save(output_path)
        logger.info(f"Файл сохранён: {output_path}")

    def prepare_table_data(rows):
        """Подготовка данных для таблицы"""
        return [
            {
                "table_number": row[0],  # Табельный номер
                "surname_name_patronymic": row[1],  # ФИО
                "profession": row[2],  # Профессия
                "percent": row[3]  # Сумма выплаты
            }
            for row in rows
        ]

    def property_parsing():
        """Парсинг данных"""
        try:
            conn, cursor = opening_the_database()
            excel_path = f"data/{year}/input/{data_mounts}/130.xlsx"
            # Открываем выбор файла Excel для чтения данных
            workbook = load_workbook(filename=excel_path)
            sheet = workbook.active

            # Создаем таблицу в базе данных, если она еще не существует
            cursor.execute(
                'CREATE TABLE IF NOT EXISTS data (table_number, surname_name_patronymic, profession, percent)')

            cursor.execute('DELETE FROM data')
            conn.commit()  # сохранить изменения

            # Считываем данные из колонок и вставляем их в базу данных
            for row in sheet.iter_rows(min_row=6, max_row=110, values_only=True):
                try:
                    # Проверяем, что строка содержит достаточно данных
                    if len(row) > 6:  # Убедимся, что в строке есть хотя бы 12 столбцов
                        table_number = row[0]  # table_number - табельный номер,
                        surname_name_patronymic = row[1]  # surname_name_patronymic - фамилия,
                        profession = row[3]  # profession - профессия,
                        percent = row[5]  # percent - процент

                        profession = full_name_of_professions.get(profession,
                                                                  profession)  # Получаем полное название профессии

                        logger.info(
                            f"Вставка: {table_number}, {surname_name_patronymic}, {profession}, {percent}"
                        )

                        # Вставляем данные в таблицу
                        cursor.execute('INSERT INTO data VALUES (?, ?, ?, ?)',
                                       (table_number, surname_name_patronymic, profession, percent))
                    else:
                        logger.warning(f"Строка пустая или неполная: {row}")
                except Exception as e:
                    logger.error(f"Ошибка при обработке строки {row}: {e}")

            # Сохраняем изменения в базе данных и закрываем соединение
            conn.commit()
            conn.close()

        except FileNotFoundError:
            logger.error(f"Excel не найден: {excel_path}")
            return

        except Exception as e:
            logger.exception(e)

    property_parsing()
    record_data_salary_downtime_week()
