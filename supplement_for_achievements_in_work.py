from docxtpl import DocxTemplate
from loguru import logger
from openpyxl import load_workbook

from database import opening_the_database, get_data_from_db
from full_name_of_professions import full_name_of_professions


def supplement_for_achievements_in_work(data_mounts, file_dog, year="2025"):
    """
    Доплата за высокие достижения в работе
    :param data_mounts: месяц, например "08"
    :param file_dog: шаблон docx
    :param year: год (по умолчанию 2025)
    """

    def record_data_salary_downtime_week():
        """Заполнение приказа"""
        doc = DocxTemplate(f"data/sample/{file_dog}")

        rows = get_data_from_db()
        table_data = prepare_table_data(rows)

        context = {
            "data_mounts": f" {data_mounts} ",
            "table_data": table_data
        }

        output_path = f"data/{year}/output/{data_mounts}/{file_dog}"
        doc.render(context)
        doc.save(output_path)
        logger.info(f"Файл сохранён: {output_path}")

    def prepare_table_data(rows):
        """Подготовка данных для таблицы"""
        return [
            {
                "table_number": row[0],
                "surname_name_patronymic": row[1],
                "profession": row[2],
                "percent": row[3],
            }
            for row in rows
        ]

    def property_parsing():
        """Парсинг Excel и запись в БД"""
        try:
            conn, cursor = opening_the_database()
            excel_path = f"data/{year}/input/{data_mounts}/124.xlsx"

            workbook = load_workbook(filename=excel_path)
            sheet = workbook.active

            cursor.execute(
                "CREATE TABLE IF NOT EXISTS data (table_number, surname_name_patronymic, profession, percent)"
            )
            cursor.execute("DELETE FROM data")
            conn.commit()

            for row in sheet.iter_rows(min_row=3, max_row=23, values_only=True):
                try:
                    if len(row) > 11:
                        table_number = row[1]
                        surname_name_patronymic = row[2]
                        profession = row[5]
                        percent = row[11]

                        profession = full_name_of_professions.get(profession, profession)

                        logger.info(
                            f"Вставка: {table_number}, {surname_name_patronymic}, {profession}, {percent}"
                        )

                        cursor.execute(
                            "INSERT INTO data VALUES (?, ?, ?, ?)",
                            (table_number, surname_name_patronymic, profession, percent),
                        )
                    else:
                        logger.warning(f"Строка пустая или неполная: {row}")
                except Exception as e:
                    logger.error(f"Ошибка при обработке строки {row}: {e}")

            conn.commit()
            conn.close()

        except FileNotFoundError:
            logger.error(f"Excel не найден: {excel_path}")
            return
        except Exception as e:
            logger.exception(e)

    property_parsing()
    record_data_salary_downtime_week()
